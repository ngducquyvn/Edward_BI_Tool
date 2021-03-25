### Step 1:  Đọc file mapping

from openpyxl import load_workbook
import os

wb_Opxl = load_workbook('Mapping.xlsx')

# Đọc sheet Tenplate
Temp_arr = []
ws_Template_Opxl = wb_Opxl["Template"]
max_rw_Template_Opxl = ws_Template_Opxl.max_row + 1
Temp_start_col = 0 # 'E' column index
Temp_end_col = 4 # 'L' column index

for i in range(1, max_rw_Template_Opxl):
	row = [cell.value for cell in ws_Template_Opxl[i][Temp_start_col:Temp_end_col+1]]
	Temp_arr.append(row)




# Đọc sheet Report
Report_arr = []
ws_Report_Opxl = wb_Opxl["Report"]
max_rw_Report_Opxl = ws_Report_Opxl.max_row + 1
Report_start_col = 0 # 'E' column index
Report_end_col = 5 # 'L' column index

for i in range(1, max_rw_Report_Opxl):
	row = [cell.value for cell in ws_Report_Opxl[i][Report_start_col:Report_end_col+1]]
	Report_arr.append(row)
	#print(row) # list of cell values of this roww



# Tạo file Mapping
Mapping_arr = []
for i in Report_arr:
	for u in Temp_arr:
		if (i[0]==u[0]):
			Mapping_arr.append(u+i)
#print(Mapping_arr[0])
#print(Mapping_arr[1])

# Đóng file
wb_Opxl.close()





### Step 2: Write đồng thời Save As file report
from editpyxl import Workbook

wb = Workbook()

#source_filename = r'TTR-CTĐ-SLGD-ORG.xlsm'
source_filename =""
output_filename =""


for i in Mapping_arr:
	if i[0] != "Template":
		source_filename =  i[1]
		output_filename =  os.path.join(i[8], i[7])
		#print(source_filename)

		wb.open(source_filename)

		ws_par = wb[i[2]]
		ws_par.cell(i[3]).value = i[4]

		wb.save(output_filename)
		wb.close()


