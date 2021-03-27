### Step 1:  Đọc file mapping

from openpyxl import load_workbook
import os
import shutil

Generate_Reports_file= 'Generate-Reports.xlsx'
wb_Opxl = load_workbook(Generate_Reports_file)

print("Loaded: " + Generate_Reports_file)

# Đọc sheet Tenplate
Temp_arr = []
ws_Template_Opxl = wb_Opxl["Template"]
max_rw_Template_Opxl = ws_Template_Opxl.max_row + 1
Temp_start_col = 0 # 'A' column index
Temp_end_col = 3 # 'E' column index


for i in range(1, max_rw_Template_Opxl):
	row = [cell.value for cell in ws_Template_Opxl[i][Temp_start_col:Temp_end_col+1]]
	Temp_arr.append(row)
print("Done: Temp_arr")

# Đọc sheet Report
Report_arr = []
ws_Report_Opxl = wb_Opxl["Report"]
max_rw_Report_Opxl = ws_Report_Opxl.max_row + 1
Report_start_col = 0 # 'A' column index
Report_end_col = 5 # 'F' column index

for i in range(1, max_rw_Report_Opxl):
	row = [cell.value for cell in ws_Report_Opxl[i][Report_start_col:Report_end_col+1]]
	Report_arr.append(row)
print("Done: Report_arr")
#print(Report_arr)


# Tạo mảng 2D Mapping
Mapping_arr = []
for i in Report_arr:
	for u in Temp_arr:
		if (i[0]==u[0]):
			Mapping_arr.append(u+i)
#print(Mapping_arr[0])
#print(Mapping_arr[1])

print("Done: 2D Array Mapping")

# Đóng file
wb_Opxl.close()
print("Closed: " + Generate_Reports_file)

## Nhân bản báo cáo

for i in Report_arr:
	if "Template" not in i[1]:
		src_dir= i[1]
		dst_dir= os.path.join(i[3], i[2])
		shutil.copy(src_dir,dst_dir)
		#print(dst_dir)


### Step 2: Write đồng thời Save As file report
from editpyxl import Workbook
''
wb = Workbook()
out_filename =""

print("*** Start generate reports ***")

# Số lượng báo cáo
len_Report_arr =len(Report_arr)-1

Cnt = 0
for r in Report_arr:
	for i in Mapping_arr:
		# r[2] và i[6] = tên báo cáo
		if (r[2] == i[6]) and ("Template" not in i[0]):
			Template_Alias = i[0]
			Sheet_dest = i[1]
			Cell_dest = i[2]
			Value_dest = i[3]
			#Template ALIAS [4]
			#Full_Path_Template =i[5]
			Report_nm = i[6]
			Short_Path_Report =i[7]
			PD_CODE = i[8]	
			Alter_value = i[9]

			out_filename =  os.path.join(Short_Path_Report, Report_nm)

			#print(out_filename)

			wb.open(out_filename)

			ws_par = wb[Sheet_dest]

			if "số lượng giao dịch" in Report_nm:
				ws_par.cell('A13').value = "Details_SLGD"
			elif "Doanh số" in Report_nm:
				ws_par.cell('A13').value = "Details_DoanhSo"
			elif "Năng suất" in Report_nm:
				ws_par.cell('A13').value = "Details_NangSuat"

			if Value_dest == "PD_CODE":
					ws_par.cell(Cell_dest).value = PD_CODE
					#print(PD_CODE)

			# Nếu Value_dest có chứa @L thì lower string nhưng ko có chữ "Báo cáo"
			elif "@L" in Value_dest:
					
				# s1 = tên chỉ tiêu
				s1 = Report_nm.replace("Báo cáo ", "").lower()
				s1 = s1.replace(".xlsm", "")

				# s2 = Tên chỉ tiêu hoàn chỉnh: thay thế @L bằng tên chỉ tiêu s1
				s2 = Value_dest.replace("@L", s1)
					
				#print (s2)

				# Nếu Value_dest có @L tại index 0 thì Upper kí tự đầu của s2
				if Value_dest.index("@L") == 0:
					s2_first = s2[0]

					ws_par.cell(Cell_dest).value = s2.replace(s2_first, s2_first.upper())
						
					#print ("@L0" + " | " + s2.replace(s2_first, s2_first.upper()))
					#print(s2.replace(s2_first, s2_first.upper()))
				else:
					ws_par.cell(Cell_dest).value = s2
					#print ("@L"+ " | " +s2)

			# Nếu có @U thì upper s1 nhưng ko có chữ "Báo cáo"		
			elif "@U" in Value_dest:
				# s1 = tên báo cáo
				s1 = Report_nm.replace("Báo cáo ", "").upper()
				s1 = s1.replace(".xlsm", "")

				# s2 = Tên chỉ tiêu hoàn chỉnh: thay thế @U bằng tên chỉ tiêu s1
				s2 = Value_dest.replace("@U", s1)

				ws_par.cell(Cell_dest).value = s2
				#print(s2)
				#print("@U" + " | " + s2)

			# @RU khác @U ở chỗ là nó lấy cả chữ Báo cáo trong string Report_nm
			elif "@RU" in Value_dest:
				# Upper tên báo cáo
				s1 = Report_nm.replace(".xlsm", "").upper()
				# 
				s2 = Value_dest.replace("@RU",s1).upper()
					
				ws_par.cell(Cell_dest).value = s2 + " "

				#print("@RU" + " | " + Value_dest.replace("@RU",s1).upper() + " ")
				#print(s2 + " ")
			elif "@Alter" in Value_dest:
				ws_par.cell(Cell_dest).value = Alter_value
				#print("@Alter" + " | " + Alter_value)
				#print(Alter_value)
			else:
				ws_par.cell(Cell_dest).value = Value_dest
				#print("Value_dest | " + Value_dest)
				#print(Value_dest)
					
		wb.save(out_filename)
		wb.close()

	if "Template" not in r[0]:
		Cnt +=1
		#print ('{}{}'.format(s, i) Cnt+out_filename)  


		print ("Done: " + repr(Cnt) + "/" + repr(len_Report_arr) + " | " + out_filename)

''' NOTE:

Thư viện Editpyxl có 1 số hạn chế về việc write như sau:
1. File đầu vào và đầu ra là phải 1, đã test với các file báo cáo thực tế, các file đơn giản hơn có thể không cần.
--> Dùng File template để copy/paste ra các báo cáo con với tên tương ứng rồi sau đó mới input/write.
2. Các cell trong file đầu vào phải có 1 giá trị, các ô liền kề nhau nên có các giá trị khác nhau,
nếu không write mặc dù success nhưng giá trị có thể bị sai lệch.
3. wb.save() và wb.close() nên cùng trong 1 session, hạn chế đóng mở đan xen nhiều session.

'''
