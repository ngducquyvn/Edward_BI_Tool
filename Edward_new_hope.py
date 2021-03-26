### Step 1:  Đọc file mapping

from openpyxl import load_workbook
import os
import shutil

wb_Opxl = load_workbook('D:\\Desktop\\Bao cao\\Mapping - Copy.xlsx')

# Đọc sheet Tenplate
Temp_arr = []
ws_Template_Opxl = wb_Opxl["Template"]
max_rw_Template_Opxl = ws_Template_Opxl.max_row + 1
Temp_start_col = 0 # 'A' column index
Temp_end_col = 3 # 'D' column index

for i in range(1, max_rw_Template_Opxl):
	row = [cell.value for cell in ws_Template_Opxl[i][Temp_start_col:Temp_end_col+1]]
	Temp_arr.append(row)




# Đọc sheet Report
Report_arr = []
ws_Report_Opxl = wb_Opxl["Report"]
max_rw_Report_Opxl = ws_Report_Opxl.max_row + 1
Report_start_col = 0 # 'E' column index
Report_end_col = 5 # 'L' column index

for i in range(1, max_rw_Report_Opxl):
	row = [cell.value for cell in ws_Report_Opxl[i][Report_start_col:Report_end_col+1]]
	Report_arr.append(row)
	#print(row) # list of cell values of this roww
#print (Report_arr[0])
print (Report_arr[1])

# Tạo file Mapping
Mapping_arr = []
for i in Report_arr:
	for u in Temp_arr:
		if (i[0]==u[0]):
			Mapping_arr.append(u+i)
#print(Mapping_arr[0])
print(Mapping_arr[1])

# Đóng file
wb_Opxl.close()

## Nhân bản báo cáo

for i in Report_arr:
	if "Template" not in i[1]:
		src_dir= i[1]
		dst_dir= os.path.join(i[3], i[2])
		shutil.copy(src_dir,dst_dir)
		#print(dst_dir)



### Step 2: Write đồng thời Save As file report
from editpyxl import Workbook

wb = Workbook()

#source_filename = r'TTR-CTĐ-SLGD-ORG.xlsm'
source_filename =""
out_filename =""

for u in Report_arr:
	for i in Mapping_arr:
		if u[2] ==i[6]:
			if "Template" not in i[0]:
				source_filename = os.path.join(i[7], i[6])
				out_filename = os.path.join("D:\\Desktop\\Bao cao\\", i[6])

				

				wb.open(source_filename)

				ws_par = wb[i[1]]
				ws_par.cell(i[2]).value = i[3]
				print(source_filename + " | " + i[3])
				
		wb.save(source_filename)
		wb.save(out_filename)
		wb.close()


